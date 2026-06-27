import csv
import io
from datetime import date

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.models.user import User
from app.services.dashboard_service import compute_kpis, get_attrition_trends
from app.services.employee_service import list_employees_query
from app.services.risk_service import high_risk_employees, risk_by_department, risk_summary

REPORT_TEMPLATE_NAMES = {
    "1": "Executive Summary",
    "2": "Attrition Analysis",
    "3": "Department Comparison",
    "4": "Cost Impact Report",
    "5": "Risk Distribution",
}


def _resolve_report_name(
    *,
    template_id: str | None = None,
    report_type: str | None = None,
    report_name: str | None = None,
) -> str:
    template_name = REPORT_TEMPLATE_NAMES.get(template_id or "", report_name or "Custom Report")
    if report_name:
        return report_name
    if report_type:
        return report_type.replace("_", " ").title()
    return template_name


def _export_filename(template_name: str, extension: str) -> str:
    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in template_name)[:40]
    return f"{safe_name or 'report'}_export.{extension}"


def _section_included(sections: list[str] | None, section: str) -> bool:
    if not sections:
        return True
    normalized_sections = {s.strip().lower() for s in sections}
    needle = section.lower()
    return needle in normalized_sections or any(needle in s for s in normalized_sections)


def _collect_report_sections(
    db: Session,
    user: User,
    *,
    sections: list[str] | None = None,
) -> list[dict]:
    report_sections: list[dict] = []

    if _section_included(sections, "Executive Summary") or _section_included(sections, "summary"):
        kpis = compute_kpis(db, user)
        report_sections.append(
            {
                "title": "Executive Summary",
                "headers": ["Metric", "Value"],
                "rows": [
                    ["Total Employees", kpis["totalEmployees"]],
                    ["Active Employees", kpis["activeEmployees"]],
                    ["Attrition Rate (%)", kpis["attritionRate"]],
                    ["Retention Rate (%)", kpis["retentionRate"]],
                    ["Avg Tenure (years)", kpis["avgTenure"]],
                    ["Avg Satisfaction", kpis["avgSatisfaction"]],
                    ["Avg Performance", kpis["avgPerformance"]],
                ],
            }
        )

    if _section_included(sections, "Risk Analysis") or _section_included(sections, "risk"):
        summary = risk_summary(db, user)
        report_sections.append(
            {
                "title": "Risk Analysis",
                "headers": ["Risk Level", "Count", "Percentage"],
                "rows": [
                    ["Low", summary["lowRisk"], f"{summary['lowRiskPct']}%"],
                    ["Medium", summary["mediumRisk"], f"{summary['mediumRiskPct']}%"],
                    ["High", summary["highRisk"], f"{summary['highRiskPct']}%"],
                    ["Avg Risk Score", summary["avgRiskScore"], ""],
                ],
            }
        )

    if _section_included(sections, "Department Breakdown") or _section_included(sections, "department"):
        departments = risk_by_department(db, user)
        report_sections.append(
            {
                "title": "Department Breakdown",
                "headers": ["Department", "Head Count", "Attrition Rate (%)", "Avg Satisfaction", "Avg Risk"],
                "rows": [
                    [
                        d["fullName"],
                        d["employeeCount"],
                        d["attritionRate"],
                        d["satisfaction"],
                        d["avgRisk"],
                    ]
                    for d in departments
                ],
            }
        )

    if _section_included(sections, "Recommendations") or _section_included(sections, "high risk"):
        at_risk = high_risk_employees(db, user, limit=25)
        report_sections.append(
            {
                "title": "High Risk Employees",
                "headers": ["Name", "Department", "Position", "Risk Level", "Attrition Probability (%)"],
                "rows": [
                    [
                        f"{e['firstName']} {e['lastName']}",
                        e.get("department", ""),
                        e.get("position", ""),
                        e.get("attritionRisk", ""),
                        e.get("attritionProbability", ""),
                    ]
                    for e in at_risk
                ],
            }
        )

    return report_sections


def _write_csv_section(writer: csv.writer, title: str, headers: list[str], rows: list[list]) -> None:
    writer.writerow([])
    writer.writerow([title])
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)


def export_report_csv(
    db: Session,
    user: User,
    *,
    template_id: str | None = None,
    report_type: str | None = None,
    report_name: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    sections: list[str] | None = None,
) -> tuple[str, str]:
    template_name = _resolve_report_name(
        template_id=template_id,
        report_type=report_type,
        report_name=report_name,
    )

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["HR Report Export"])
    writer.writerow(["Report", template_name])
    if start_date:
        writer.writerow(["Start Date", start_date.isoformat()])
    if end_date:
        writer.writerow(["End Date", end_date.isoformat()])
    writer.writerow(["Generated By", user.email])

    for section in _collect_report_sections(db, user, sections=sections):
        _write_csv_section(writer, section["title"], section["headers"], section["rows"])

    return buffer.getvalue(), _export_filename(template_name, "csv")


def export_report_xlsx(
    db: Session,
    user: User,
    *,
    template_id: str | None = None,
    report_type: str | None = None,
    report_name: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    sections: list[str] | None = None,
) -> tuple[bytes, str]:
    template_name = _resolve_report_name(
        template_id=template_id,
        report_type=report_type,
        report_name=report_name,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"
    bold = Font(bold=True)

    worksheet.append(["HR Report Export"])
    worksheet["A1"].font = bold
    worksheet.append(["Report", template_name])
    if start_date:
        worksheet.append(["Start Date", start_date.isoformat()])
    if end_date:
        worksheet.append(["End Date", end_date.isoformat()])
    worksheet.append(["Generated By", user.email])
    worksheet.append([])

    for section in _collect_report_sections(db, user, sections=sections):
        title_row = worksheet.max_row + 1
        worksheet.cell(row=title_row, column=1, value=section["title"]).font = bold
        worksheet.append(section["headers"])
        header_row = worksheet.max_row
        for cell in worksheet[header_row]:
            cell.font = bold
        for row in section["rows"]:
            worksheet.append(row)
        worksheet.append([])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), _export_filename(template_name, "xlsx")


def export_report_pdf(
    db: Session,
    user: User,
    *,
    template_id: str | None = None,
    report_type: str | None = None,
    report_name: str | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    sections: list[str] | None = None,
) -> tuple[bytes, str]:
    template_name = _resolve_report_name(
        template_id=template_id,
        report_type=report_type,
        report_name=report_name,
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=18,
        spaceAfter=12,
    )
    heading_style = ParagraphStyle(
        "SectionHeading",
        parent=styles["Heading2"],
        fontSize=13,
        spaceBefore=14,
        spaceAfter=8,
    )
    meta_style = styles["Normal"]

    story: list = [Paragraph(f"HR Report: {template_name}", title_style)]
    meta_lines = [f"Generated by: {user.email}"]
    if start_date:
        meta_lines.append(f"Start date: {start_date.isoformat()}")
    if end_date:
        meta_lines.append(f"End date: {end_date.isoformat()}")
    story.append(Paragraph("<br/>".join(meta_lines), meta_style))
    story.append(Spacer(1, 0.2 * inch))

    for section in _collect_report_sections(db, user, sections=sections):
        story.append(Paragraph(section["title"], heading_style))
        table_data = [section["headers"]] + [[str(cell) for cell in row] for row in section["rows"]]
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f7fa")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    return buffer.getvalue(), _export_filename(template_name, "pdf")


def build_report_preview(db: Session, user: User) -> dict:
    kpis = compute_kpis(db, user)
    departments = risk_by_department(db, user)
    return {
        "kpis": kpis,
        "departments": departments[:5],
        "pages": 1,
    }


def get_turnover_cost_breakdown(db: Session, user: User) -> dict:
    employees = list_employees_query(db, user).all()
    kpis = compute_kpis(db, user)
    avg_salary = sum(float(employee.salary) for employee in employees) / len(employees) if employees else 1_000_000
    components = [
        ("Recruitment", 0.25),
        ("Training", 0.20),
        ("Lost Productivity", 0.35),
        ("Administrative", 0.10),
        ("Onboarding", 0.10),
    ]
    total_cost = round(avg_salary * sum(weight for _, weight in components))
    breakdown = []
    for category, weight in components:
        cost = round(avg_salary * weight)
        breakdown.append(
            {
                "category": category,
                "cost": cost,
                "percentage": round(weight * 100),
            }
        )
    monthly_turnover = max(1, round(kpis["highRiskCount"] / 15))
    return {
        "breakdown": breakdown,
        "avgTurnoverCost": total_cost,
        "monthlyTurnover": monthly_turnover,
        "annualImpact": total_cost * monthly_turnover * 12,
        "currency": "RWF",
    }


def get_factor_correlations(db: Session, user: User) -> list[dict]:
    employees = list_employees_query(db, user).all()
    if len(employees) < 20:
        return _static_correlations()

    probabilities = np.array([employee.attrition_probability for employee in employees], dtype=float)
    feature_specs = [
        ("Satisfaction", np.array([employee.satisfaction_score for employee in employees], dtype=float), "Higher satisfaction tends to reduce attrition risk"),
        ("Tenure", np.array([employee.years_at_company for employee in employees], dtype=float), "Longer tenure can reduce short-term attrition risk"),
        ("Overtime", np.array([employee.overtime_hours for employee in employees], dtype=float), "Higher overtime correlates with attrition risk"),
        ("Performance", np.array([employee.performance_score for employee in employees], dtype=float), "Performance patterns vs predicted attrition"),
        ("Work-Life Balance", np.array([employee.work_life_balance for employee in employees], dtype=float), "Better balance tends to reduce attrition risk"),
        ("Training Hours", np.array([employee.training_hours for employee in employees], dtype=float), "Development investment vs attrition risk"),
    ]
    results: list[dict] = []
    for factor, values, description in feature_specs:
        if np.std(values) == 0 or np.std(probabilities) == 0:
            correlation = 0.0
        else:
            correlation = float(np.corrcoef(values, probabilities)[0, 1])
        results.append(
            {
                "factor": factor,
                "correlation": round(correlation, 2),
                "description": description,
            }
        )
    results.sort(key=lambda item: abs(item["correlation"]), reverse=True)
    return results


def get_attrition_forecast(db: Session, user: User) -> list[dict]:
    trends = get_attrition_trends()
    kpis = compute_kpis(db, user)
    forecast_rows: list[dict] = []
    for row in trends:
        forecast_rows.append(
            {
                "month": row["month"],
                "actual": row["actual"],
                "predicted": row["predicted"],
                "forecast": None,
            }
        )

    current_rate = float(kpis["attritionRate"])
    future_months = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    projected = [round(current_rate * factor, 1) for factor in [0.95, 0.92, 0.98, 0.94, 0.90, 0.88]]
    for month, value in zip(future_months, projected):
        forecast_rows.append(
            {
                "month": month,
                "actual": None,
                "predicted": None,
                "forecast": value,
            }
        )
    return forecast_rows


def _static_correlations() -> list[dict]:
    return [
        {"factor": "Satisfaction", "correlation": -0.72, "description": "Strong negative correlation with attrition"},
        {"factor": "Tenure", "correlation": -0.45, "description": "Moderate negative correlation"},
        {"factor": "Overtime", "correlation": 0.58, "description": "Positive correlation with attrition risk"},
    ]
