import csv
import io
import random
from datetime import date, datetime
from decimal import Decimal
from uuid import UUID

from fastapi import HTTPException, status
from pydantic import ValidationError
from sqlalchemy import func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, joinedload

from openpyxl import Workbook, load_workbook

from app.core.permissions import UserRole
from app.models.organization import (
    AttritionRisk,
    CompensationRecord,
    Department,
    Employee,
    EmployeeCertification,
    EmployeeSkill,
    EmploymentEvent,
    EmployeeStatus,
)
from app.models.user import User
from app.schemas.employee import EmployeeCreate, EmployeeOut, EmployeeStats, EmployeeUpdate

EMPLOYEE_CSV_HEADERS = [
    "employeeId",
    "firstName",
    "lastName",
    "email",
    "department",
    "position",
    "hireDate",
    "salary",
    "currency",
    "payFrequency",
    "age",
    "gender",
    "yearsAtCompany",
    "performanceScore",
    "satisfactionScore",
    "workLifeBalance",
    "lastPromotionYears",
    "trainingHours",
    "overtimeHours",
    "status",
    "attritionRisk",
    "attritionProbability",
]

EMPLOYEE_IMPORT_HEADERS = [
    "employeeId",
    "firstName",
    "lastName",
    "email",
    "department",
    "position",
    "hireDate",
    "salary",
    "currency",
    "payFrequency",
    "age",
    "gender",
    "yearsAtCompany",
    "performanceScore",
    "satisfactionScore",
    "workLifeBalance",
    "lastPromotionYears",
    "trainingHours",
    "overtimeHours",
]

_HEADER_ALIASES: dict[str, str] = {
    "employeeid": "employeeId",
    "employee_id": "employeeId",
    "firstname": "firstName",
    "first_name": "firstName",
    "lastname": "lastName",
    "last_name": "lastName",
    "email": "email",
    "department": "department",
    "position": "position",
    "hiredate": "hireDate",
    "hire_date": "hireDate",
    "salary": "salary",
    "currency": "currency",
    "payfrequency": "payFrequency",
    "pay_frequency": "payFrequency",
    "age": "age",
    "gender": "gender",
    "yearsatcompany": "yearsAtCompany",
    "years_at_company": "yearsAtCompany",
    "performancescore": "performanceScore",
    "performance_score": "performanceScore",
    "satisfactionscore": "satisfactionScore",
    "satisfaction_score": "satisfactionScore",
    "worklifebalance": "workLifeBalance",
    "work_life_balance": "workLifeBalance",
    "lastpromotionyears": "lastPromotionYears",
    "last_promotion_years": "lastPromotionYears",
    "traininghours": "trainingHours",
    "training_hours": "trainingHours",
    "overtimehours": "overtimeHours",
    "overtime_hours": "overtimeHours",
    "status": "status",
    "attritionrisk": "attritionRisk",
    "attrition_risk": "attritionRisk",
    "attritionprobability": "attritionProbability",
    "attrition_probability": "attritionProbability",
}


def _normalize_header(header: str) -> str | None:
    key = header.strip().replace(" ", "").replace("-", "_").lower()
    return _HEADER_ALIASES.get(key)


DEPARTMENT_ALIASES: dict[str, str] = {
    "customer service": "Customer Support",
    "customer support": "Customer Support",
    "it support": "Customer Support",
    "hr": "Human Resources",
    "human resources": "Human Resources",
    "r&d": "Research & Development",
    "research and development": "Research & Development",
    "research & development": "Research & Development",
}

_INTEGER_IMPORT_FIELDS = {"age", "yearsAtCompany", "lastPromotionYears", "trainingHours", "overtimeHours"}


def _years_at_company_from_hire_date(hire_date: date, reference: date | None = None) -> int:
    reference = reference or date.today()
    years = reference.year - hire_date.year
    if (reference.month, reference.day) < (hire_date.month, hire_date.day):
        years -= 1
    return max(0, years)


def _parse_import_hire_date(value: str) -> date | None:
    cleaned = value.strip()
    if not cleaned:
        return None
    try:
        return date.fromisoformat(cleaned)
    except ValueError:
        for fmt in ("%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(cleaned, fmt).date()
            except ValueError:
                continue
    return None


def _apply_hire_date_tenure(payload_data: dict[str, str]) -> dict[str, str]:
    hire_date = _parse_import_hire_date(payload_data.get("hireDate", ""))
    if hire_date:
        payload_data["yearsAtCompany"] = str(_years_at_company_from_hire_date(hire_date))
    return payload_data


def _coerce_import_field(header: str, value: str) -> str:
    if not value:
        return value
    if header in _INTEGER_IMPORT_FIELDS:
        try:
            return str(int(round(float(value.replace(",", "")))))
        except ValueError:
            return value
    if header == "department":
        mapped = DEPARTMENT_ALIASES.get(value.strip().lower())
        if mapped:
            return mapped
    return value


def _normalize_import_row(payload_data: dict[str, str]) -> dict[str, str]:
    return {
        header: _coerce_import_field(header, value)
        for header, value in payload_data.items()
    }


def _calculate_risk(
    satisfaction: float,
    years: int,
    promotion_years: int,
    overtime: int,
    work_life: float,
) -> tuple[AttritionRisk, int]:
    probability = 0.1
    if satisfaction < 7:
        probability += 0.2
    if years < 2:
        probability += 0.15
    if promotion_years > 3:
        probability += 0.15
    if overtime > 20:
        probability += 0.1
    if work_life < 6.5:
        probability += 0.15
    probability = min(0.95, max(0.05, probability + random.uniform(-0.1, 0.1)))
    pct = int(round(probability * 100))
    if pct < 30:
        risk = AttritionRisk.LOW
    elif pct < 60:
        risk = AttritionRisk.MEDIUM
    else:
        risk = AttritionRisk.HIGH
    return risk, pct


def serialize_employee(employee: Employee) -> dict:
    return EmployeeOut(
        id=employee.id,
        employee_id=employee.employee_id,
        first_name=employee.first_name,
        last_name=employee.last_name,
        email=employee.email,
        department=employee.department.name if employee.department else "",
        position=employee.position,
        hire_date=employee.hire_date,
        salary=employee.salary,
        currency=employee.currency,
        pay_frequency=employee.pay_frequency,
        age=employee.age,
        gender=employee.gender,
        years_at_company=employee.years_at_company,
        performance_score=employee.performance_score,
        satisfaction_score=employee.satisfaction_score,
        work_life_balance=employee.work_life_balance,
        last_promotion_years=employee.last_promotion_years,
        training_hours=employee.training_hours,
        overtime_hours=employee.overtime_hours,
        attrition_risk=employee.attrition_risk,
        attrition_probability=employee.attrition_probability,
        status=employee.status,
        avatar_url=employee.avatar_url,
    ).model_dump(by_alias=True)


def get_department_scope(user: User) -> UUID | None:
    if user.role == UserRole.DEPARTMENT_HEAD:
        return user.department_id
    return None


def list_employees_query(db: Session, user: User):
    query = (
        db.query(Employee)
        .options(joinedload(Employee.department))
        .filter(Employee.deleted_at.is_(None))
    )
    dept_scope = get_department_scope(user)
    if dept_scope:
        query = query.filter(Employee.department_id == dept_scope)
    return query


def list_employees(
    db: Session,
    user: User,
    *,
    search: str | None = None,
    department: str | None = None,
    risk: str | None = None,
    status_filter: str | None = None,
    page: int = 1,
    limit: int = 20,
    sort: str = "lastName",
) -> tuple[list[Employee], int]:
    query = list_employees_query(db, user)

    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(
                Employee.first_name.ilike(pattern),
                Employee.last_name.ilike(pattern),
                Employee.email.ilike(pattern),
                Employee.employee_id.ilike(pattern),
            )
        )
    if department:
        query = query.join(Department).filter(Department.name.ilike(f"%{department}%"))
    if risk:
        query = query.filter(Employee.attrition_risk == risk)
    if status_filter:
        query = query.filter(Employee.status == status_filter)

    total = query.count()
    sort_map = {
        "lastName": Employee.last_name.asc(),
        "firstName": Employee.first_name.asc(),
        "department": Employee.department_id.asc(),
        "risk": Employee.attrition_probability.desc(),
    }
    query = query.order_by(sort_map.get(sort, Employee.last_name.asc()))
    items = query.offset((page - 1) * limit).limit(limit).all()
    return items, total


def get_employee(db: Session, user: User, employee_id: UUID) -> Employee:
    employee = (
        list_employees_query(db, user)
        .filter(Employee.id == employee_id)
        .first()
    )
    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"code": "NOT_FOUND", "message": "Employee not found"},
        )
    return employee


def resolve_department(db: Session, department_name: str) -> Department:
    department = db.query(Department).filter(Department.name == department_name).first()
    if not department:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "INVALID_DEPARTMENT", "message": f"Department '{department_name}' not found"},
        )
    return department


def _department_from_cache(department_by_name: dict[str, Department], department_name: str) -> Department | None:
    department = department_by_name.get(department_name)
    if department:
        return department
    canonical = DEPARTMENT_ALIASES.get(department_name.strip().lower())
    if canonical:
        return department_by_name.get(canonical)
    return None


def _validate_new_employee(db: Session, user: User, payload: EmployeeCreate) -> Department:
    department = resolve_department(db, payload.department)
    if get_department_scope(user) and department.id != user.department_id:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={"code": "FORBIDDEN", "message": "Cannot create employee outside your department"},
        )
    existing_email = (
        db.query(Employee)
        .filter(Employee.email == payload.email, Employee.deleted_at.is_(None))
        .first()
    )
    if existing_email:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "DUPLICATE_EMAIL", "message": f"Email already exists: {payload.email}"},
        )
    if payload.employee_id:
        existing_code = db.query(Employee).filter(Employee.employee_id == payload.employee_id).first()
        if existing_code:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={"code": "DUPLICATE_EMPLOYEE_ID", "message": f"Employee ID already exists: {payload.employee_id}"},
            )
    return department


def _add_employee_to_session(
    db: Session,
    department: Department,
    payload: EmployeeCreate,
    *,
    employee_code: str | None = None,
) -> Employee:
    if not employee_code:
        count = db.query(func.count(Employee.id)).scalar() or 0
        employee_code = payload.employee_id or f"EMP{str(count + 1001).zfill(5)}"
    risk, probability = _calculate_risk(
        payload.satisfaction_score,
        payload.years_at_company,
        payload.last_promotion_years,
        payload.overtime_hours,
        payload.work_life_balance,
    )
    employee = Employee(
        employee_id=employee_code,
        first_name=payload.first_name,
        last_name=payload.last_name,
        email=payload.email,
        department_id=department.id,
        position=payload.position,
        hire_date=payload.hire_date,
        salary=payload.salary,
        currency=payload.currency,
        pay_frequency=payload.pay_frequency,
        age=payload.age,
        gender=payload.gender,
        years_at_company=payload.years_at_company,
        performance_score=payload.performance_score,
        satisfaction_score=payload.satisfaction_score,
        work_life_balance=payload.work_life_balance,
        last_promotion_years=payload.last_promotion_years,
        training_hours=payload.training_hours,
        overtime_hours=payload.overtime_hours,
        attrition_risk=risk,
        attrition_probability=probability,
        status=payload.status,
    )
    db.add(employee)
    db.flush()
    db.add(
        EmploymentEvent(
            employee_id=employee.id,
            event_type="hire",
            title="Joined company",
            description=f"Started as {employee.position} in {department.name}",
            event_date=employee.hire_date,
        )
    )
    db.add(
        CompensationRecord(
            employee_id=employee.id,
            base_salary=employee.salary,
            bonus=Decimal(str(float(employee.salary) * 0.1)),
            stock_options=max(100, int(float(employee.salary) / 1000)),
            effective_date=employee.hire_date,
        )
    )
    return employee


def create_employee(db: Session, user: User, payload: EmployeeCreate) -> Employee:
    department = _validate_new_employee(db, user, payload)
    employee = _add_employee_to_session(db, department, payload)
    db.commit()
    db.refresh(employee)
    return get_employee(db, user, employee.id)


def update_employee(db: Session, user: User, employee_id: UUID, payload: EmployeeUpdate) -> Employee:
    employee = get_employee(db, user, employee_id)
    data = payload.model_dump(exclude_unset=True)
    if "department" in data:
        department = resolve_department(db, data.pop("department"))
        employee.department_id = department.id
    for key, value in data.items():
        setattr(employee, key, value)
    risk, probability = _calculate_risk(
        employee.satisfaction_score,
        employee.years_at_company,
        employee.last_promotion_years,
        employee.overtime_hours,
        employee.work_life_balance,
    )
    employee.attrition_risk = risk
    employee.attrition_probability = probability
    db.commit()
    db.refresh(employee)
    return employee


def employee_stats(db: Session, user: User) -> dict:
    query = list_employees_query(db, user)
    total = query.count()
    active = query.filter(Employee.status == EmployeeStatus.ACTIVE).count()
    on_leave = query.filter(Employee.status == EmployeeStatus.ON_LEAVE).count()
    inactive = query.filter(Employee.status == EmployeeStatus.INACTIVE).count()
    high_risk = query.filter(Employee.attrition_risk == AttritionRisk.HIGH).count()
    medium_risk = query.filter(Employee.attrition_risk == AttritionRisk.MEDIUM).count()
    low_risk = query.filter(Employee.attrition_risk == AttritionRisk.LOW).count()
    avg_performance = db.query(func.avg(Employee.performance_score)).filter(
        Employee.deleted_at.is_(None)
    ).scalar() or 0
    avg_satisfaction = db.query(func.avg(Employee.satisfaction_score)).filter(
        Employee.deleted_at.is_(None)
    ).scalar() or 0
    return EmployeeStats(
        total=total,
        active=active,
        on_leave=on_leave,
        inactive=inactive,
        high_risk=high_risk,
        medium_risk=medium_risk,
        low_risk=low_risk,
        avg_performance=round(float(avg_performance), 1),
        avg_satisfaction=round(float(avg_satisfaction), 1),
    ).model_dump(by_alias=True)


def _collect_employee_export_rows(
    db: Session,
    user: User,
    *,
    search: str | None = None,
    department: str | None = None,
    risk: str | None = None,
    status_filter: str | None = None,
) -> list[dict]:
    items, _ = list_employees(
        db,
        user,
        search=search,
        department=department,
        risk=risk,
        status_filter=status_filter,
        page=1,
        limit=100_000,
    )
    return [
        {header: serialize_employee(employee).get(header, "") for header in EMPLOYEE_CSV_HEADERS}
        for employee in items
    ]


def export_employees_csv(
    db: Session,
    user: User,
    *,
    search: str | None = None,
    department: str | None = None,
    risk: str | None = None,
    status_filter: str | None = None,
) -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=EMPLOYEE_CSV_HEADERS)
    writer.writeheader()
    for row in _collect_employee_export_rows(
        db,
        user,
        search=search,
        department=department,
        risk=risk,
        status_filter=status_filter,
    ):
        writer.writerow(row)
    return buffer.getvalue()


def export_employees_xlsx(
    db: Session,
    user: User,
    *,
    search: str | None = None,
    department: str | None = None,
    risk: str | None = None,
    status_filter: str | None = None,
) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Employees"
    worksheet.append(EMPLOYEE_CSV_HEADERS)
    for row in _collect_employee_export_rows(
        db,
        user,
        search=search,
        department=department,
        risk=risk,
        status_filter=status_filter,
    ):
        worksheet.append([row.get(header, "") for header in EMPLOYEE_CSV_HEADERS])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _import_template_sample_row() -> dict[str, str]:
    hire_date = date(2024, 1, 15)
    return {
        "employeeId": "EMP01001",
        "firstName": "Jane",
        "lastName": "Doe",
        "email": "jane.doe@company.com",
        "department": "Engineering",
        "position": "Software Engineer",
        "hireDate": hire_date.isoformat(),
        "salary": "500000",
        "currency": "RWF",
        "payFrequency": "monthly",
        "age": "28",
        "gender": "female",
        "yearsAtCompany": str(_years_at_company_from_hire_date(hire_date)),
        "performanceScore": "7.5",
        "satisfactionScore": "7.0",
        "workLifeBalance": "7.0",
        "lastPromotionYears": "1",
        "trainingHours": "10",
        "overtimeHours": "5",
    }


def employee_import_template_csv() -> str:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=EMPLOYEE_IMPORT_HEADERS)
    writer.writeheader()
    writer.writerow(_import_template_sample_row())
    return buffer.getvalue()


def employee_import_template_xlsx() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Import Template"
    worksheet.append(EMPLOYEE_IMPORT_HEADERS)
    sample = _import_template_sample_row()
    worksheet.append([sample.get(header, "") for header in EMPLOYEE_IMPORT_HEADERS])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _cell_to_import_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def _parse_xlsx_rows(raw: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        row_iter = worksheet.iter_rows(values_only=True)
        header_row = next(row_iter, None)
        if not header_row:
            return []
        headers = [str(header).strip() if header is not None else "" for header in header_row]
        rows: list[dict[str, str]] = []
        for values in row_iter:
            if not any(value is not None and _cell_to_import_str(value) for value in values):
                continue
            row: dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header or index >= len(values):
                    continue
                row[header] = _cell_to_import_str(values[index])
            rows.append(row)
        return rows
    finally:
        workbook.close()


def _validate_import_headers(fieldnames: list[str | None]) -> None:
    normalized_fieldnames = [_normalize_header(name or "") for name in fieldnames]
    if not any(normalized_fieldnames):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "INVALID_FILE", "message": "Spreadsheet headers are not recognized"},
        )


def _import_employee_rows(db: Session, user: User, rows: list[dict[str, str]]) -> dict:
    existing_emails = {
        email
        for (email,) in db.query(Employee.email).filter(Employee.deleted_at.is_(None)).all()
    }
    existing_codes = {code for (code,) in db.query(Employee.employee_id).all()}
    department_by_name = {dept.name: dept for dept in db.query(Department).all()}

    created = 0
    skipped = 0
    errors: list[dict] = []
    seen_employee_ids: set[str] = set()
    base_count = db.query(func.count(Employee.id)).scalar() or 0

    for row_number, row in enumerate(rows, start=2):
        if not any(value and str(value).strip() for value in row.values()):
            continue

        payload_data: dict[str, str] = {}
        for raw_header, value in row.items():
            header = _normalize_header(raw_header or "")
            if header and header not in {"attritionRisk", "attritionProbability", "status"}:
                payload_data[header] = str(value).strip() if value is not None else ""

        payload_data = _normalize_import_row(payload_data)
        payload_data = _apply_hire_date_tenure(payload_data)

        if not payload_data.get("email"):
            errors.append({"row": row_number, "message": "Email is required"})
            skipped += 1
            continue

        email = payload_data["email"]
        if email in existing_emails:
            errors.append({"row": row_number, "message": f"Email already exists: {email}"})
            skipped += 1
            continue

        employee_code = payload_data.get("employeeId", "").strip()
        if employee_code:
            if employee_code in seen_employee_ids:
                errors.append({"row": row_number, "message": f"Duplicate employee ID in file: {employee_code}"})
                skipped += 1
                continue
            seen_employee_ids.add(employee_code)
            if employee_code in existing_codes:
                errors.append({"row": row_number, "message": f"Employee ID already exists: {employee_code}"})
                skipped += 1
                continue
        else:
            payload_data.pop("employeeId", None)

        try:
            payload = EmployeeCreate.model_validate(payload_data)
            department = _department_from_cache(department_by_name, payload.department)
            if not department:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail={"code": "INVALID_DEPARTMENT", "message": f"Department '{payload.department}' not found"},
                )
            if get_department_scope(user) and department.id != user.department_id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail={"code": "FORBIDDEN", "message": "Cannot create employee outside your department"},
                )
            auto_code = None
            if not payload.employee_id:
                auto_code = f"EMP{str(base_count + created + 1001).zfill(5)}"
            _add_employee_to_session(db, department, payload, employee_code=auto_code or payload.employee_id)
            existing_emails.add(email)
            if employee_code:
                existing_codes.add(employee_code)
            created += 1
        except ValidationError as exc:
            err = exc.errors()[0]
            field = err["loc"][-1] if err.get("loc") else "field"
            message = f"{field}: {err['msg']}"
            errors.append({"row": row_number, "message": message})
            skipped += 1
        except HTTPException as exc:
            detail = exc.detail if isinstance(exc.detail, dict) else {"message": str(exc.detail)}
            errors.append({"row": row_number, "message": detail.get("message", "Import failed")})
            skipped += 1
        except IntegrityError:
            errors.append({"row": row_number, "message": f"Could not create employee: {email}"})
            skipped += 1

    return {"created": created, "skipped": skipped, "errors": errors}


def _csv_reader(content: str) -> csv.DictReader:
    try:
        dialect = csv.Sniffer().sniff(content[:4096], delimiters=",\t")
    except csv.Error:
        dialect = csv.excel
    return csv.DictReader(io.StringIO(content), dialect=dialect)


def import_employees_csv(db: Session, user: User, content: str) -> dict:
    reader = _csv_reader(content)
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "INVALID_CSV", "message": "CSV file is empty or missing headers"},
        )
    _validate_import_headers(list(reader.fieldnames))
    return _import_employee_rows(db, user, list(reader))


def import_employees_xlsx(db: Session, user: User, raw: bytes) -> dict:
    rows = _parse_xlsx_rows(raw)
    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"code": "INVALID_FILE", "message": "Excel file is empty or missing headers"},
        )
    _validate_import_headers(list(rows[0].keys()))
    return _import_employee_rows(db, user, rows)
